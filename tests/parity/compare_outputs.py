"""Output-vs-output parity: diff the ORIGINAL Streamlit ranked workbook against
the CURRENT ranked workbook for each client.

The route-optimizer-factory optimizer is byte-identical to the dashboard
backend's, so current-output == factory-output; diffing original-vs-current
validates the extraction end-to-end when you have output workbooks (not the raw
penetration-report inputs).

Client workbooks are proprietary and git-ignored. Drop them under
`tests/parity/fixtures/` and list the pairs in
`tests/parity/outputs_manifest.json` (copy `outputs_manifest.example.json`):

    { "pairs": [ { "name": "C&C Myers",
                   "original": "cc-myers/original.xlsx",
                   "current":  "cc-myers/current.xlsx" } ] }

Run:  .venv/bin/python tests/parity/compare_outputs.py [manifest]

Known-benign differences (reported but not counted as failures):
  - sheet ORDER (current puts 'Ranked Geocodes' first) — intended
  - '$'-prefixed income/home-value bracket labels in 'Dashboard Data' — shipped feature
  - 'Summary' audit rows in a different order but identical per-label values — cosmetic
Exit 0 = parity (ignoring known-benign); 1 = real differences or error.
"""
from __future__ import annotations

import json
import math
import os
import sys

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
FIXTURES = os.path.join(HERE, "fixtures")
DEFAULT_MANIFEST = os.path.join(HERE, "outputs_manifest.json")
TOL = 1e-9


def resolve(path):
    return path if os.path.isabs(path) else os.path.join(FIXTURES, path)


def read_sheet(path, sheet):
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return None
    rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    wb.close()
    return rows


def close(a, b):
    if isinstance(a, (int, float)) and isinstance(b, (int, float)) \
            and not isinstance(a, bool) and not isinstance(b, bool):
        if math.isnan(a) and math.isnan(b):
            return True
        return abs(a - b) <= TOL
    return a == b


def same_modulo_dollar(a, b):
    return isinstance(a, str) and isinstance(b, str) and a.replace("$", "") == b.replace("$", "")


def compare_ranked(orig, cur):
    o = read_sheet(orig, "Ranked Geocodes")
    c = read_sheet(cur, "Ranked Geocodes")
    oh, ch = o[0], c[0]
    gi_o, gi_c = oh.index("Geocode"), ch.index("Geocode")
    od = {r[gi_o]: r for r in o[1:]}
    cd = {r[gi_c]: r for r in c[1:]}
    shared = [col for col in oh if col in ch]
    key_diff = set(od) ^ set(cd)
    mism = {}
    for g in set(od) & set(cd):
        for col in shared:
            a, b = od[g][oh.index(col)], cd[g][ch.index(col)]
            if not close(a, b):
                mism[col] = mism.get(col, 0) + 1
    order_match = [r[gi_o] for r in o[1:]] == [r[gi_c] for r in c[1:]]
    print(f"  Ranked Geocodes: {len(o)-1} rows, {len(shared)} shared cols")
    if not mism and not key_diff:
        print(f"    OK  all columns exact; ranking order identical={order_match}")
        return True
    print(f"    DIFF  key-set diff={len(key_diff)}; columns with diffs={mism}")
    return False


def compare_benign(orig, cur, sheet):
    """Returns real-diff count after excluding known-benign deltas."""
    o = read_sheet(orig, sheet)
    c = read_sheet(cur, sheet)
    if o is None or c is None:
        return 0
    real = 0
    if sheet == "Summary":
        od = {r[0]: r for r in o if r and r[0] is not None}
        cd = {r[0]: r for r in c if r and r[0] is not None}
        for k in set(od) & set(cd):
            for a, b in zip(od[k][1:], cd[k][1:]):
                if not close(a, b):
                    real += 1
        if set(od) ^ set(cd):
            real += len(set(od) ^ set(cd))
        print(f"  {sheet}: aligned-by-label real diffs={real} (row order ignored)")
    else:
        for oro, cro in zip(o, c):
            for a, b in zip(oro, cro):
                if not close(a, b) and not same_modulo_dollar(a, b):
                    real += 1
        print(f"  {sheet}: real diffs={real} ('$'-bracket labels ignored)")
    return real


def main(argv=None):
    argv = argv or sys.argv[1:]
    manifest_path = argv[0] if argv else DEFAULT_MANIFEST
    if not os.path.exists(manifest_path):
        print(f"No manifest at {manifest_path}. Copy outputs_manifest.example.json "
              f"to outputs_manifest.json and drop the client workbooks under fixtures/.")
        return 1
    with open(manifest_path, encoding="utf-8") as h:
        pairs = json.load(h).get("pairs", [])
    if not pairs:
        print("No pairs in manifest.")
        return 1

    all_ok = True
    for pair in pairs:
        print("=" * 64)
        print(pair["name"])
        try:
            ok = compare_ranked(resolve(pair["original"]), resolve(pair["current"]))
            dd = compare_benign(resolve(pair["original"]), resolve(pair["current"]), "Dashboard Data")
            sm = compare_benign(resolve(pair["original"]), resolve(pair["current"]), "Summary")
        except FileNotFoundError as e:
            print(f"  missing file: {e}")
            all_ok = False
            continue
        all_ok = all_ok and ok and dd == 0 and sm == 0
    print("=" * 64)
    print("OVERALL:", "PARITY (known-benign deltas ignored)" if all_ok else "REAL DIFFERENCES FOUND")
    return 0 if all_ok else 1


if __name__ == "__main__":
    sys.exit(main())
