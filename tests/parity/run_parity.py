"""Parity harness: prove the extracted optimizer reproduces the original tool.

Runs `cli.py` against committed client report inputs and diffs the result against
a captured baseline (`result.json` + optionally `ranked.xlsx`). This is the gate
for Phase 1 — the 4 known client pairs must match, including C&C Myers (the
0.9x Composite Score case fixed by `_effective_weights`).

Baselines are NOT committed (client data). Drop them under
`tests/parity/fixtures/<case>/` and describe each case in a manifest:

    python tests/parity/run_parity.py tests/parity/manifest.json

Manifest schema (see manifest.example.json):
    {
      "cases": [
        {
          "name": "cc-myers",
          "mode": "score",                       # "score" | "roi"
          "args": { ...cli score/roi args... },  # filePath relative to fixtures/
          "baselineResult": "cc-myers/baseline.result.json",
          "baselineXlsx":   "cc-myers/baseline.ranked.xlsx",   # optional
          "floatTolerance": 1e-6
        }
      ]
    }

Exit code 0 = all cases match; 1 = any mismatch or error.
"""
from __future__ import annotations

import argparse
import json
import math
import os
import subprocess
import sys

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
CLI = os.path.join(REPO_ROOT, "optimizer", "cli.py")
FIXTURES = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures")
PYTHON = sys.executable
DEFAULT_TOL = 1e-6


def _resolve(path: str) -> str:
    """Resolve a manifest path relative to the fixtures dir (or absolute)."""
    return path if os.path.isabs(path) else os.path.join(FIXTURES, path)


def _num_close(a, b, tol: float) -> bool:
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        if a is None or b is None:
            return a == b
        if math.isnan(a) and math.isnan(b):
            return True
        return abs(a - b) <= tol
    return a == b


def diff_value(path: str, a, b, tol: float, out: list) -> None:
    if isinstance(a, dict) and isinstance(b, dict):
        for key in sorted(set(a) | set(b)):
            if key not in a:
                out.append(f"{path}.{key}: missing in actual")
            elif key not in b:
                out.append(f"{path}.{key}: missing in baseline")
            else:
                diff_value(f"{path}.{key}", a[key], b[key], tol, out)
    elif isinstance(a, list) and isinstance(b, list):
        if len(a) != len(b):
            out.append(f"{path}: length {len(a)} != baseline {len(b)}")
            return
        for i, (x, y) in enumerate(zip(a, b)):
            diff_value(f"{path}[{i}]", x, y, tol, out)
    elif not _num_close(a, b, tol):
        out.append(f"{path}: {a!r} != baseline {b!r}")


def run_case(case: dict) -> list[str]:
    name = case["name"]
    mode = case.get("mode", "score")
    tol = case.get("floatTolerance", DEFAULT_TOL)
    args = dict(case["args"])
    if "filePath" in args:
        args["filePath"] = _resolve(args["filePath"])
    if "roiFilePath" in args:
        args["roiFilePath"] = _resolve(args["roiFilePath"])

    out_dir = os.path.join(FIXTURES, ".out", name)
    os.makedirs(out_dir, exist_ok=True)

    proc = subprocess.run(
        [PYTHON, CLI, "--mode", mode, "--out-dir", out_dir],
        input=json.dumps(args), capture_output=True, text=True,
    )
    if not proc.stdout.strip():
        return [f"cli produced no stdout (exit {proc.returncode}); stderr: {proc.stderr[:500]}"]
    actual = json.loads(proc.stdout)
    if "error" in actual:
        return [f"cli returned error: {actual['error']}"]

    problems: list[str] = []
    with open(_resolve(case["baselineResult"]), "r", encoding="utf-8") as handle:
        baseline = json.load(handle)
    # Compare the result-defining fields; ignore volatile meta like reportLabel.
    for field in ("dashboardModel", "rankedRows", "rankedColumns",
                  "mergedRows", "mergedColumns", "roiSummary"):
        if field in baseline or field in actual:
            diff_value(field, actual.get(field), baseline.get(field), tol, problems)

    if case.get("baselineXlsx"):
        problems += diff_xlsx(os.path.join(out_dir, actual["artifacts"]["xlsx"]),
                              _resolve(case["baselineXlsx"]), tol)
    return problems


def diff_xlsx(actual_path: str, baseline_path: str, tol: float) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ["openpyxl not installed; cannot diff xlsx"]
    problems: list[str] = []
    a = load_workbook(actual_path, data_only=True)
    b = load_workbook(baseline_path, data_only=True)
    for sheet in b.sheetnames:
        if sheet not in a.sheetnames:
            problems.append(f"xlsx: sheet '{sheet}' missing in actual")
            continue
        sa, sb = a[sheet], b[sheet]
        for row in range(1, sb.max_row + 1):
            for col in range(1, sb.max_column + 1):
                va = sa.cell(row=row, column=col).value
                vb = sb.cell(row=row, column=col).value
                if not _num_close(va, vb, tol):
                    problems.append(f"xlsx[{sheet}]!R{row}C{col}: {va!r} != baseline {vb!r}")
    return problems


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description="Route Optimizer parity harness")
    parser.add_argument("manifest", help="Path to a parity manifest JSON")
    parser.add_argument("--max-report", type=int, default=20,
                        help="Max mismatch lines to print per case")
    args = parser.parse_args(argv)

    with open(args.manifest, "r", encoding="utf-8") as handle:
        manifest = json.load(handle)
    cases = manifest.get("cases", [])
    if not cases:
        print("No cases in manifest. Add client pairs under tests/parity/fixtures/ "
              "and list them in the manifest (see manifest.example.json).")
        return 1

    failed = 0
    for case in cases:
        name = case["name"]
        try:
            problems = run_case(case)
        except FileNotFoundError as err:
            problems = [f"missing file: {err}"]
        if problems:
            failed += 1
            print(f"FAIL  {name}  ({len(problems)} mismatch(es))")
            for line in problems[:args.max_report]:
                print(f"      {line}")
            if len(problems) > args.max_report:
                print(f"      … +{len(problems) - args.max_report} more")
        else:
            print(f"PASS  {name}")

    print(f"\n{len(cases) - failed}/{len(cases)} cases matched.")
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
