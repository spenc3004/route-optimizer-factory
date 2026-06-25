"""End-to-end parity: run the EXTRACTED cli.py on raw penetration reports and
match the ORIGINAL Streamlit Composite Scores.

The exact UI jobSpec (preset/weights/driver) isn't recorded in the output, but
the known client runs used 'Dynamic (from file)' profile mode (ideals derived
from the file → reproducible). So we brute-force the small config space
(runnable preset x available driver, with preset-default weights) and find the
combo whose Composite Scores match the original exactly. A match both proves the
engine on real inputs AND recovers the config used. Fail filters don't affect
Composite Score, so the comparison is keyed by Geocode (order/Status independent).

Manifest-driven (same file as compare_outputs.py); each pair needs `original`,
`input`, and `category`:

    .venv/bin/python tests/parity/recover_and_verify.py [manifest]

Client data is proprietary and git-ignored; drop it under fixtures/.
"""
from __future__ import annotations

import json
import os
import subprocess
import sys

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(os.path.dirname(HERE))
CLI = os.path.join(REPO, "optimizer", "cli.py")
PY = sys.executable
FIXTURES = os.path.join(HERE, "fixtures")
DEFAULT_MANIFEST = os.path.join(HERE, "outputs_manifest.json")

CONFIG = json.load(open(os.path.join(REPO, "contract", "config.json")))
BYCAT = {c["name"]: c for c in CONFIG["categories"]}
FIXED_ONLY = {"Home SRVCS Acquisition (No History)", "Auto Acquisition (No History)"}
EXACT = 1e-9


def resolve(p):
    return p if os.path.isabs(p) else os.path.join(FIXTURES, p)


def og_scores(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    rows = list(wb["Ranked Geocodes"].iter_rows(values_only=True))
    wb.close()
    h = rows[0]
    gi, ci = h.index("Geocode"), h.index("Composite Score")
    return {r[gi]: r[ci] for r in rows[1:]}


def input_columns(path):
    import pandas as pd
    return set(pd.read_excel(path, nrows=1).columns)


def run_score(category, preset, driver, input_path, out_dir):
    args = {
        "category": category, "preset": preset, "filePath": input_path,
        "originalName": os.path.basename(input_path),
        "weights": BYCAT[category]["defaultWeightsByPreset"][preset],
        "profile": {"mode": "Dynamic (from file)", "driver": driver},
    }
    proc = subprocess.run([PY, CLI, "--mode", "score", "--out-dir", out_dir],
                          input=json.dumps(args), capture_output=True, text=True)
    out = json.loads(proc.stdout) if proc.stdout.strip() else {}
    if "error" in out:
        return None
    return {r["Geocode"]: r["Composite Score"] for r in out["rankedRows"]}


def max_diff(a, b):
    common = set(a) & set(b)
    md = 0.0
    for g in common:
        if a[g] is not None and b[g] is not None:
            md = max(md, abs(a[g] - b[g]))
    return md, len(common)


def verify_pair(pair):
    name = pair["name"]
    category = pair["category"]
    cat = BYCAT[category]
    input_path = resolve(pair["input"])
    target = og_scores(resolve(pair["original"]))
    cols = input_columns(input_path)
    drivers = [d for d in cat["drivers"] if d in cols]
    presets = [p for p in cat["presets"] if p not in FIXED_ONLY]
    out_dir = os.path.join(FIXTURES, ".out", name)
    os.makedirs(out_dir, exist_ok=True)

    best = None
    for preset in presets:
        for driver in drivers:
            got = run_score(category, preset, driver, input_path, out_dir)
            if got is None:
                continue
            md, n = max_diff(got, target)
            if best is None or md < best[0]:
                best = (md, preset, driver, n)
            if md <= EXACT:
                return best
    return best


def main(argv=None):
    argv = argv or sys.argv[1:]
    manifest_path = argv[0] if argv else DEFAULT_MANIFEST
    if not os.path.exists(manifest_path):
        print(f"No manifest at {manifest_path}. See outputs_manifest.example.json; "
              f"each pair needs original + input + category.")
        return 1
    pairs = [p for p in json.load(open(manifest_path)).get("pairs", [])
             if p.get("input") and p.get("category")]
    if not pairs:
        print("No pairs with input+category in manifest.")
        return 1

    overall = True
    for pair in pairs:
        print("=" * 64)
        print(pair["name"], f"({pair['category']})")
        best = verify_pair(pair)
        if best is None:
            print("  no runnable combo"); overall = False; continue
        md, preset, driver, n = best
        verdict = "EXACT MATCH" if md <= EXACT else ("close" if md <= 1e-6 else "NO MATCH")
        print(f"  recovered config: preset={preset!r} driver={driver!r}")
        print(f"  Composite Score max|Δ| over {n} geocodes = {md:.3g}  -> {verdict}")
        overall = overall and md <= EXACT
    print("=" * 64)
    print("OVERALL:", "END-TO-END PARITY (raw input -> original output)" if overall
          else "NOT ALL EXACT (see above)")
    return 0 if overall else 1


if __name__ == "__main__":
    sys.exit(main())
