from __future__ import annotations

import math
import os
import re

import numpy as np
import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText, Text
from openpyxl.chart.title import Title
from openpyxl.drawing.image import Image
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.spreadsheet_drawing import (
    AnchorMarker,
    OneCellAnchor,
    TwoCellAnchor,
)
from openpyxl.drawing.text import (
    CharacterProperties,
    Font as DrawingFont,
    Paragraph,
    ParagraphProperties,
    RegularTextRun,
)
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# --- Palette (ARGB) ---------------------------------------------------------
NAVY = "FF002B42"          # banners, title band, table headers, dark bars
BLUE = "FF3091D0"          # profile header row, blue bars
ORANGE = "FFFF8C00"        # accent stripe, orange bars (#FF8C00, matches the brand orange)
WHITE = "FFFFFFFF"
GRAY_TEXT = "FF686766"     # subtitle + muted labels
BORDER_GRAY = "FFC9D2D8"   # thin cell borders
ALT_ROW = "FFEAEAE8"       # zebra striping in the profile table

# --- Typography / number formats --------------------------------------------
FONT_NAME = "Arial"
INT_FMT = "#,##0"
CURRENCY_FMT = '"$"#,##0'

# --- Layout -----------------------------------------------------------------
LEFT_PAD = 1               # blank gutter column(s) on the far left
FIRST_COL = 1 + LEFT_PAD   # B — content starts after the padding gutter
LAST_COL = FIRST_COL + 17  # 18 content columns wide (six 3-wide KPI cards)
DATA_SHEET = "Dashboard Data"
SHEET_NAME = "Dashboard"
TOP_N = 10      # top-N ZIPs per Geographic Performance chart

# --- Print setup ------------------------------------------------------------
PRINT_LAST_ROW = 92          # bottom of the printable dashboard (B1:S92)
PAPER_LETTER = 1             # OOXML paper-size code for US Letter
# Excel "Narrow" margin preset (inches).
PRINT_MARGINS = dict(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3)

# --- Logo -------------------------------------------------------------------
LOGO_PATH = os.path.join(os.path.dirname(__file__), os.pardir, "assets",
                         "MailSharkLogoBadge.png")
LOGO_HEIGHT_PX = 124         # left-aligned, vertically centered in the title band
LOGO_INSET_PX = 51           # gap from the band's left edge

# --- Column widths ----------------------------------------------------------
GUTTER_COL_WIDTH = 8.5       # blank left-padding column A
FIRST_CONTENT_WIDTH = 9.16   # column B (narrower so the KPI cards line up)
CONTENT_COL_WIDTH = 8.43     # columns C–S (the body of the dashboard)

# --- Title sizing -----------------------------------------------------------
TITLE_MAX_PT = 28            # font size for normal-length report labels
TITLE_MIN_PT = 16           # smallest line-1 size before truncating with an ellipsis
SUBTITLE_MAX_PT = 16         # line-2 size when line 1 is at TITLE_MAX_PT
SUBTITLE_RATIO = SUBTITLE_MAX_PT / TITLE_MAX_PT
SUBTITLE_MIN_PT = 10         # legibility floor for line 2
TITLE_CHAR_W = 0.62          # avg glyph width as a fraction of point size (Arial bold caps)
TITLE_INDENT_PX = 9          # approx pixels per Excel indent unit
TITLE_FIT_SAFETY = 0.92      # target this fraction of the available width

# --- Chart placement / sizing -----------------------------------------------
GEO_CHART_ANCHORS = {
    "Selected":    (0, 15, 5, 28, 357120, 186120),
    "House Count": (6, 15, 11, 28, 357480, 186120),
    "Total Spend": (12, 15, 17, 28, 357120, 186120),
}
CARRIER_CHART_ANCHORS = {  # mirrors the ZIP charts, 18 rows lower
    "Selected":    (0, 33, 5, 46, 357120, 186120),
    "House Count": (6, 33, 11, 46, 357480, 186120),
    "Total Spend": (12, 33, 17, 46, 357120, 186120),
}
REVENUE_CHART_ANCHORS = {
    "$ Income":     (0, 51, 8, 65, 212040, 67320),
    "$ Home Value": (9, 51, 17, 65, 211680, 67320),
}
CUSTOMER_CHART_ANCHORS = {  # mirrors the revenue charts, 18 rows lower
    "$ Income":     (0, 69, 8, 83, 212040, 67320),
    "$ Home Value": (9, 69, 17, 83, 211680, 67320),
}
GEO_GAP_WIDTH = 40
REVENUE_GAP_WIDTH = 150
CHART_TITLE_SIZE = 1600  # 16pt, in OOXML hundredths-of-a-point — the next standard
                         # step below 18pt that keeps the longest title ("House Count
                         # by Carrier Route") on one line in the middle-row charts.
GRIDLINE_COLOR = "B3B3B3"
GRIDLINE_WIDTH = 9525    # ~0.75pt in EMU
AXIS_CURRENCY_FMT = '"$"#,##0,"K"'
CHART_HORIZONTAL_SHIFT_EMU = 100000

# --- Ideal Profile + Explanation panels -------------------------------------
SUMMARY_BANNER_ROW = 86      # below the Carrier + Revenue + Customer histogram sections
SUMMARY_FIRST_ROW = 87
SUMMARY_MAX_ROW = 95
# Offsets from FIRST_COL: profile panel 8 cols, 2-col gap, explanation 8 cols.
PROFILE_PANEL_COLS = (FIRST_COL, FIRST_COL + 7)
EXPLANATION_PANEL_COLS = (FIRST_COL + 10, FIRST_COL + 17)
PROFILE_LABEL_COLS = (FIRST_COL, FIRST_COL + 3)
PROFILE_VALUE_COLS = (FIRST_COL + 4, FIRST_COL + 7)


# === Small styling helpers ==================================================

def _fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color)


def _font(size: float, *, bold=False, italic=False, color=NAVY) -> Font:
    return Font(name=FONT_NAME, size=size, bold=bold, italic=italic, color=color)


def _align(horizontal=None, vertical="center", wrap=False, indent=0) -> Alignment:
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap, indent=indent)


def _outline(ws, r1: int, c1: int, r2: int, c2: int, color: str = BORDER_GRAY) -> None:
    side = Side(style="thin", color=color)
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            cur = ws.cell(row=row, column=col).border
            ws.cell(row=row, column=col).border = Border(
                top=side if row == r1 else cur.top,
                bottom=side if row == r2 else cur.bottom,
                left=side if col == c1 else cur.left,
                right=side if col == c2 else cur.right,
            )


def _write(ws, row, col, value, *, font=None, fill=None, align=None,
           number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if align is not None:
        cell.alignment = align
    if number_format is not None:
        cell.number_format = number_format
    return cell


def _merge(ws, r1, c1, r2, c2, value=None, *, font=None, fill=None, align=None,
           number_format=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    if fill is not None:
        # Fill every cell so the merged block paints solidly (openpyxl quirk).
        for row in range(r1, r2 + 1):
            for col in range(c1, c2 + 1):
                ws.cell(row=row, column=col).fill = fill
    return _write(ws, r1, c1, value, font=font, align=align,
                  number_format=number_format)


def _banner(ws, row, text, *, c1=FIRST_COL, c2=LAST_COL):
    _merge(ws, row, c1, row, c2, text,
           font=_font(11, bold=True, color=WHITE), fill=_fill(NAVY),
           align=_align(horizontal="left"))


# === Data model =============================================================

def _col(df: pd.DataFrame, name: str) -> pd.Series | None:
    return df[name] if name in df.columns else None


def _zip_label(value) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    if isinstance(value, bool):  # bool is an int subclass
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    stripped = re.fullmatch(r"(\d+)\.0+", text)
    return stripped.group(1) if stripped else text


def _zip_aggregates(df: pd.DataFrame) -> pd.DataFrame:
    if "Zip Code" not in df.columns:
        return pd.DataFrame(columns=["ZIP", "City", "Selected", "House Count", "Total Spend"])

    work = pd.DataFrame({"ZIP": df["Zip Code"]})
    work["City"] = _col(df, "City") if _col(df, "City") is not None else ""
    for src, dst in (("Selected", "Selected"),
                     ("House Count", "House Count"),
                     ("$ Total Spend", "Total Spend")):
        series = _col(df, src)
        work[dst] = pd.to_numeric(series, errors="coerce") if series is not None else 0.0

    grouped = (
        work.groupby("ZIP", dropna=True)
        .agg(City=("City", "first"),
             Selected=("Selected", "sum"),
             **{"House Count": ("House Count", "sum")},
             **{"Total Spend": ("Total Spend", "sum")})
        .reset_index()
    )
    grouped["Selected"] = grouped["Selected"].fillna(0)
    grouped["House Count"] = grouped["House Count"].fillna(0)
    grouped["Total Spend"] = grouped["Total Spend"].fillna(0)
    return grouped.sort_values("Selected", ascending=False).reset_index(drop=True)


def _carrier_routes(df: pd.DataFrame) -> pd.DataFrame:
    if "Geocode" not in df.columns:
        return pd.DataFrame(columns=["Route", "Selected", "House Count", "Total Spend"])

    out = pd.DataFrame({"Route": df["Geocode"]})
    for src, dst in (("Selected", "Selected"),
                     ("House Count", "House Count"),
                     ("$ Total Spend", "Total Spend")):
        series = _col(df, src)
        out[dst] = pd.to_numeric(series, errors="coerce") if series is not None else 0.0
    for col in ("Selected", "House Count", "Total Spend"):
        out[col] = out[col].fillna(0)
    return out


def _kpis(df: pd.DataFrame, zips: pd.DataFrame) -> list[tuple[str, object, str]]:
    return [
        ("HOUSE COUNT", int(zips["House Count"].sum()), INT_FMT),
        ("TOTAL SPEND", float(zips["Total Spend"].sum()), CURRENCY_FMT),
        ("ZIP CODES", int(zips["ZIP"].nunique()), INT_FMT),
        ("GEOCODES", int(len(df)), INT_FMT),
        ("CITIES", int(df["City"].nunique()) if "City" in df.columns else 0, INT_FMT),
    ]


INCOME_BRACKETS = [
    (0, 39_999, "0-39,999"),
    (40_000, 74_999, "40,000-74,999"),
    (75_000, 149_999, "75,000-149,999"),
    (150_000, 174_999, "150,000-174,999"),
    (175_000, 249_999, "175,000-249,999"),
    (250_000, None, "250,000+"),
]
HOME_VALUE_BRACKETS = [
    (0, 149_999, "0-149,999"),
    (150_000, 249_999, "150,000-249,999"),
    (250_000, 399_999, "250,000-399,999"),
    (400_000, 749_999, "400,000-749,999"),
    (750_000, 999_999, "750,000-999,999"),
    (1_000_000, None, "1,000,000+"),
]


def _revenue_col(df: pd.DataFrame) -> str | None:
    for candidate in ("$ Total Spend", "Overall Revenue", "Revenue"):
        if candidate in df.columns:
            return candidate
    return None


def _dollar_bracket(label: str) -> str:
    return re.sub(r"(\d[\d,]*)", r"$\1", label)


def _driver_label(driver: str) -> str:
    label = str(driver).strip().lstrip("$").strip()
    return label[:-1].strip() if label.endswith("%") else label


def build_profile_explanation(profile_mode: str | None,
                              profile_driver: str | None = None) -> str:
    if profile_mode == "Fixed Standard":
        return ("The customer profile ideals are standardized benchmark values used by "
                "Mail Shark and can be adjusted as needed to fit specific campaign objectives.")
    if profile_mode == "Dynamic (from file)":
        if profile_driver:
            return ("The customer profile ideals were calculated using the top-performing "
                    f"25% of geocodes ranked by {_driver_label(profile_driver)}. The values "
                    "shown to the left represent the average characteristics of those geocodes.")
        return ("The customer profile ideals were calculated using the top-performing 25% of "
                "geocodes in the uploaded file. The values shown to the left represent their "
                "average characteristics.")
    return ("The customer profile ideals summarize the target characteristics used to score "
            "each geocode.")


# === On-screen dashboard model (parity with the Excel sheet) ================

def _json_number(value):
    if value is None:
        return None
    if isinstance(value, (int,)) and not isinstance(value, bool):
        return value
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return int(number) if number.is_integer() else number


def _top_n_series(frame: pd.DataFrame, label_col: str, metric: str) -> list[dict]:
    if frame.empty or metric not in frame.columns or label_col not in frame.columns:
        return []
    top = frame.sort_values(metric, ascending=False).head(TOP_N)
    return [{"zip": _zip_label(label), "value": _json_number(value)}
            for label, value in zip(top[label_col], top[metric])]


def _top_zip_series(zips: pd.DataFrame, metric: str) -> list[dict]:
    return _top_n_series(zips, "ZIP", metric)


def _bracket_series(df: pd.DataFrame, value_col: str, total_col: str,
                    brackets, dollar_labels: bool = False) -> list[dict]:
    from utilities.helpers import bracket_revenue_summary

    if value_col not in df.columns or total_col not in df.columns:
        return []
    summary = bracket_revenue_summary(df, value_col, total_col, brackets)
    series = []
    for bracket, value in zip(summary["Bracket"], summary["Total Revenue"]):
        label = _dollar_bracket(bracket) if dollar_labels else bracket
        series.append({"bracket": label, "value": _json_number(value)})
    return series


def _composite_histogram(df: pd.DataFrame, bins: int = 20) -> list[dict]:
    if "Composite Score" not in df.columns:
        return []
    scores = pd.to_numeric(df["Composite Score"], errors="coerce").dropna()
    if scores.empty:
        return []
    counts, edges = np.histogram(scores, bins=bins)
    return [{"bin": f"{edges[i]:.2f}-{edges[i + 1]:.2f}", "count": int(counts[i])}
            for i in range(len(counts))]


def build_dashboard_model(working_df: pd.DataFrame,
                          base_summary_dict: dict | None = None,
                          profile_mode: str | None = None,
                          profile_driver: str | None = None) -> dict:

    base_summary_dict = base_summary_dict or {"Note": [], "Value": []}
    zips = _zip_aggregates(working_df)
    routes = _carrier_routes(working_df)
    kpi_keys = ["houseCount", "totalSpend", "zipCount", "geocodeCount", "cityCount"]
    kpis = {key: _json_number(value)
            for key, (_, value, _) in zip(kpi_keys, _kpis(working_df, zips))}
    revenue_col = _revenue_col(working_df)

    return {
        "kpis": kpis,
        "zipCharts": {
            "selected": _top_zip_series(zips, "Selected"),
            "houseCount": _top_zip_series(zips, "House Count"),
            "totalSpend": _top_zip_series(zips, "Total Spend"),
        },
        "carrierRoutes": {
            "selected": _top_n_series(routes, "Route", "Selected"),
            "houseCount": _top_n_series(routes, "Route", "House Count"),
            "totalSpend": _top_n_series(routes, "Route", "Total Spend"),
        },
        "revenueByIncomeBracket": (
            _bracket_series(working_df, "$ Income", revenue_col, INCOME_BRACKETS,
                            dollar_labels=True)
            if revenue_col else []),
        "revenueByHomeValueBracket": (
            _bracket_series(working_df, "$ Home Value", revenue_col, HOME_VALUE_BRACKETS,
                            dollar_labels=True)
            if revenue_col else []),
        "customersByIncomeBracket": _bracket_series(
            working_df, "$ Income", "Customer Count", INCOME_BRACKETS, dollar_labels=True),
        "customersByHomeValueBracket": _bracket_series(
            working_df, "$ Home Value", "Customer Count", HOME_VALUE_BRACKETS, dollar_labels=True),
        "compositeHistogram": _composite_histogram(working_df),
        "profile": {
            "explanation": build_profile_explanation(profile_mode, profile_driver),
            "rows": [{"note": note, "value": value}
                     for note, value in zip(base_summary_dict.get("Note", []),
                                            base_summary_dict.get("Value", []))],
        },
    }


# === Companion data sheet (chart series) ====================================

class _DataSheet:

    def __init__(self, ws):
        self.ws = ws
        ws.sheet_state = "hidden"
        _write(ws, 1, 1, "Dashboard chart data — generated automatically; do not edit.",
               font=_font(9, italic=True, color=GRAY_TEXT))
        self.row = 3

    def block(self, title, categories, values, value_fmt=INT_FMT):
        ws = self.ws
        _write(ws, self.row, 1, title, font=_font(9, bold=True, color=NAVY))
        header = self.row + 1
        _write(ws, header, 1, "Category", font=_font(9, bold=True))
        _write(ws, header, 2, "Value", font=_font(9, bold=True))
        first = header + 1
        for i, (cat, val) in enumerate(zip(categories, values)):
            _write(ws, first + i, 1, cat, number_format="@")
            _write(ws, first + i, 2, val, number_format=value_fmt)
        last = first + len(values) - 1
        cats = Reference(ws, min_col=1, min_row=first, max_row=last)
        vals = Reference(ws, min_col=2, min_row=first, max_row=last)
        self.row = last + 3
        return cats, vals


def _two_cell_anchor(spec: tuple[int, int, int, int, int, int]) -> TwoCellAnchor:
    from_col, from_row, to_col, to_row, to_col_off, to_row_off = spec
    shift = CHART_HORIZONTAL_SHIFT_EMU
    return TwoCellAnchor(
        _from=AnchorMarker(col=from_col + LEFT_PAD, row=from_row, colOff=shift, rowOff=0),
        to=AnchorMarker(col=to_col + LEFT_PAD, row=to_row, colOff=to_col_off + shift, rowOff=to_row_off),
    )


def _grey_gridlines() -> ChartLines:
    return ChartLines(spPr=GraphicalProperties(
        ln=LineProperties(solidFill=GRIDLINE_COLOR, w=GRIDLINE_WIDTH)))


def _chart_title(text: str) -> Title:
    props = CharacterProperties(latin=DrawingFont(typeface="Calibri"),
                                sz=CHART_TITLE_SIZE, b=True)
    paragraph = Paragraph(pPr=ParagraphProperties(defRPr=props),
                          r=[RegularTextRun(rPr=props, t=text)])
    title = Title(tx=Text(rich=RichText(p=[paragraph])))
    title.overlay = False
    return title


def _bar_chart(title, cats, vals, color, *, horizontal, gap_width, value_fmt=None):
    chart = BarChart()
    chart.type = "bar" if horizontal else "col"
    chart.grouping = "clustered"
    chart.title = _chart_title(title)
    chart.legend = None
    chart.gapWidth = gap_width
    chart.add_data(vals, titles_from_data=False)
    chart.set_categories(cats)
    chart.series[0].cat = AxDataSource(strRef=StrRef(f=str(cats)))

    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.majorGridlines = None
    chart.y_axis.majorGridlines = _grey_gridlines()
    chart.x_axis.tickLblSkip = 1
    chart.x_axis.tickMarkSkip = 1
    if value_fmt:
        axis_fmt = AXIS_CURRENCY_FMT if (horizontal and value_fmt == CURRENCY_FMT) else value_fmt
        chart.y_axis.numFmt = axis_fmt
        chart.y_axis.number_format = axis_fmt

    series = chart.series[0]
    series.graphicalProperties = GraphicalProperties(solidFill=color[-6:])  # chart fills want 6-digit RGB
    series.dLbls = DataLabelList(
        showVal=True, showSerName=False, showCatName=False,
        showLegendKey=False, showPercent=False, showBubbleSize=False,
    )
    if value_fmt:
        series.dLbls.numFmt = value_fmt
    return chart


# === Public entry point =====================================================

def add_dashboard_sheet(writer, working_df: pd.DataFrame,
                        base_summary_dict: dict | None = None,
                        report_label: str | None = None,
                        profile_mode: str | None = None,
                        profile_driver: str | None = None) -> None:
    
    base_summary_dict = base_summary_dict or {"Note": [], "Value": []}
    wb = writer.book

    ws = wb.create_sheet(SHEET_NAME)
    ws.sheet_view.showGridLines = False
    data = _DataSheet(wb.create_sheet(DATA_SHEET))

    zips = _zip_aggregates(working_df)

    _size_columns(ws)  # before the header so the logo can right-align off real widths
    _render_header(ws, report_label)
    _render_kpis(ws, _kpis(working_df, zips))
    _render_geographic_charts(ws, data, zips)
    _render_carrier_charts(ws, data, working_df)
    _render_revenue_charts(ws, data, working_df)
    _render_customer_charts(ws, data, working_df)
    _render_profile_summary(ws, base_summary_dict, profile_mode, profile_driver)
    _setup_print(ws)

    wb._sheets.remove(ws)
    wb._sheets.insert(1, ws)


# === Section renderers ======================================================

def _fit_title_size(text: str, available_px: float) -> tuple[str, int]:
    target = available_px * TITLE_FIT_SAFETY
    n = max(len(text), 1)
    size = TITLE_MAX_PT
    while size > TITLE_MIN_PT and n * TITLE_CHAR_W * size > target:
        size -= 1
    if n * TITLE_CHAR_W * size > target:  # still too wide at the floor → ellipsize
        budget = max(1, int(target / (TITLE_CHAR_W * size)) - 1)
        text = text[:budget].rstrip() + "…"
    return text, size


def _render_header(ws, report_label: str | None) -> None:
    label = (report_label or "").strip().upper()
    line1 = label or "EXECUTIVE TARGETING & REVENUE DASHBOARD"
    line2 = "EXECUTIVE TARGETING & REVENUE DASHBOARD" if label else ""
    text_c1 = FIRST_COL + 3  # B–D backdrop the logo; title starts at E
    text_indent = 18         # pushes the title toward the banner centre, clear of the logo

    col_px = round(CONTENT_COL_WIDTH * 7) + 5
    available_px = (LAST_COL - text_c1 + 1) * col_px - text_indent * TITLE_INDENT_PX
    line1, line1_size = _fit_title_size(line1, available_px)
    line2_size = max(SUBTITLE_MIN_PT, round(line1_size * SUBTITLE_RATIO * 2) / 2)


    _merge(ws, 1, FIRST_COL, 3, text_c1 - 1, fill=_fill(NAVY))  # logo backdrop
    _merge(ws, 1, text_c1, 1, LAST_COL, line1,
           font=_font(line1_size, bold=True, color=WHITE), fill=_fill(NAVY),
           align=_align(horizontal="left", vertical="bottom", indent=text_indent))
    _merge(ws, 2, text_c1, 2, LAST_COL, fill=_fill(NAVY))  # gap between the lines
    _merge(ws, 3, text_c1, 3, LAST_COL, line2,
           font=_font(line2_size, bold=True, color=WHITE), fill=_fill(NAVY),
           align=_align(horizontal="left", vertical="top", indent=text_indent))
    # Orange accent strip + subtitle beneath the title band.
    _merge(ws, 4, FIRST_COL, 4, LAST_COL, fill=_fill(ORANGE))
    _merge(ws, 5, FIRST_COL, 5, LAST_COL,
           "Geographic concentration, revenue distribution and ideal-profile "
           "summary  ·  Mail Shark direct-mail targeting model",
           font=_font(9, italic=True, color=GRAY_TEXT), fill=_fill(WHITE),
           align=_align(horizontal="left"))
    for row, height in ((1, 58.0), (2, 10.0), (3, 52.0), (4, 3.75),
                        (5, 15.75), (6, 15.0)):
        ws.row_dimensions[row].height = height
    _render_logo(ws)


def _render_logo(ws) -> None:
    if not os.path.exists(LOGO_PATH):
        return
    img = Image(LOGO_PATH)
    ratio = img.width / img.height
    img.height = LOGO_HEIGHT_PX
    img.width = round(LOGO_HEIGHT_PX * ratio)

    band_px = sum(round((ws.row_dimensions[r].height or 15) * 4 / 3) for r in (1, 2, 3))
    target_top_px = max(0, (band_px - img.height) // 2)
    acc, from_row, row_off_px = 0, 0, target_top_px
    for r in (1, 2, 3):
        h = round((ws.row_dimensions[r].height or 15) * 4 / 3)
        if acc + h > target_top_px:
            from_row, row_off_px = r - 1, target_top_px - acc
            break
        acc += h

    img.anchor = OneCellAnchor(
        _from=AnchorMarker(col=FIRST_COL - 1, row=from_row,
                           colOff=pixels_to_EMU(LOGO_INSET_PX), rowOff=pixels_to_EMU(row_off_px)),
        ext=XDRPositiveSize2D(pixels_to_EMU(img.width), pixels_to_EMU(img.height)),
    )
    ws.add_image(img)


def _render_kpis(ws, kpis: list[tuple[str, object, str]]) -> None:
    _banner(ws, 7, "KEY METRICS")
    ws.row_dimensions[7].height = 19.5
    ws.row_dimensions[9].height = 4.5
    ws.row_dimensions[10].height = 13.5
    ws.row_dimensions[11].height = 15.75
    ws.row_dimensions[12].height = 13.5

    n = len(kpis)
    total_cols = LAST_COL - FIRST_COL + 1
    for i, (label, value, fmt) in enumerate(kpis):
        c1 = FIRST_COL + round(i * total_cols / n)
        c2 = FIRST_COL + round((i + 1) * total_cols / n) - 1
        _merge(ws, 9, c1, 9, c2, fill=_fill(ORANGE))
        _merge(ws, 10, c1, 10, c2, label,
               font=_font(9, bold=True, color=GRAY_TEXT), fill=_fill(WHITE),
               align=_align(horizontal="left"))
        _merge(ws, 11, c1, 12, c2, value,
               font=_font(18, bold=True, color=NAVY), fill=_fill(WHITE),
               align=_align(horizontal="left"), number_format=fmt)
        _outline(ws, 9, c1, 12, c2)
    ws.row_dimensions[13].height = 15.0


def _render_geographic_charts(ws, data: _DataSheet, zips: pd.DataFrame) -> None:
    _banner(ws, 14, "GEOGRAPHIC PERFORMANCE  —  TOP 10 PERFORMING ZIP CODES")
    ws.row_dimensions[14].height = 19.5

    if zips.empty:
        _merge(ws, 16, FIRST_COL, 16, LAST_COL,
               "No ZIP-level data available in the source file.",
               font=_font(10, italic=True, color=GRAY_TEXT), fill=_fill(WHITE),
               align=_align(horizontal="left"))
        return

    specs = [
        ("Selected", "Selected by ZIP", NAVY, INT_FMT),
        ("House Count", "House Count by ZIP", BLUE, INT_FMT),
        ("Total Spend", "Total Spend by ZIP", ORANGE, CURRENCY_FMT),
    ]
    for metric, title, color, fmt in specs:
        top = (zips.sort_values(metric, ascending=False)
                   .head(TOP_N)
                   .iloc[::-1])
        cats, vals = data.block(title,
                                [_zip_label(z) for z in top["ZIP"]],
                                list(top[metric]),
                                value_fmt=fmt)
        chart = _bar_chart(title, cats, vals, color, horizontal=True,
                           gap_width=GEO_GAP_WIDTH, value_fmt=fmt)
        chart.anchor = _two_cell_anchor(GEO_CHART_ANCHORS[metric])
        ws.add_chart(chart)


def _render_carrier_charts(ws, data: _DataSheet, df: pd.DataFrame) -> None:
    _banner(ws, 32, "GEOGRAPHIC PERFORMANCE  —  TOP 10 PERFORMING CARRIER ROUTES")
    ws.row_dimensions[32].height = 19.5

    routes = _carrier_routes(df)
    if routes.empty:
        _merge(ws, 34, FIRST_COL, 34, LAST_COL,
               "No carrier-route data available in the source file.",
               font=_font(10, italic=True, color=GRAY_TEXT), fill=_fill(WHITE),
               align=_align(horizontal="left"))
        return

    specs = [
        ("Selected", "Selected by Carrier Route", NAVY, INT_FMT),
        ("House Count", "House Count by Carrier Route", BLUE, INT_FMT),
        ("Total Spend", "Total Spend by Carrier Route", ORANGE, CURRENCY_FMT),
    ]
    for metric, title, color, fmt in specs:
        top = (routes.sort_values(metric, ascending=False)
                     .head(TOP_N)
                     .iloc[::-1])
        cats, vals = data.block(title,
                                [_zip_label(r) for r in top["Route"]],
                                list(top[metric]),
                                value_fmt=fmt)
        chart = _bar_chart(title, cats, vals, color, horizontal=True,
                           gap_width=GEO_GAP_WIDTH, value_fmt=fmt)
        chart.anchor = _two_cell_anchor(CARRIER_CHART_ANCHORS[metric])
        ws.add_chart(chart)


def _render_revenue_charts(ws, data: _DataSheet, df: pd.DataFrame) -> None:
    from utilities.helpers import bracket_revenue_summary

    _banner(ws, 50, "REVENUE ANALYSIS  —  DISTRIBUTION HISTOGRAMS")
    ws.row_dimensions[50].height = 19.5

    revenue_col = _revenue_col(df)
    histo_specs = [
        ("$ Income", INCOME_BRACKETS, "Revenue by Income", BLUE),
        ("$ Home Value", HOME_VALUE_BRACKETS, "Revenue by Home Value", ORANGE),
    ]
    if revenue_col is None:
        _merge(ws, 52, FIRST_COL, 52, LAST_COL,
               "Revenue histograms unavailable — no revenue column in source data.",
               font=_font(10, italic=True, color=GRAY_TEXT), fill=_fill(WHITE),
               align=_align(horizontal="left"))
        return

    for value_col, brackets, title, color in histo_specs:
        if value_col not in df.columns:
            continue
        summary = bracket_revenue_summary(df, value_col, revenue_col, brackets)
        labels = [_dollar_bracket(b) for b in summary["Bracket"]]
        cats, vals = data.block(title, labels,
                                list(summary["Total Revenue"]),
                                value_fmt=CURRENCY_FMT)
        chart = _bar_chart(title, cats, vals, color, horizontal=False,
                           gap_width=REVENUE_GAP_WIDTH, value_fmt=CURRENCY_FMT)
        chart.anchor = _two_cell_anchor(REVENUE_CHART_ANCHORS[value_col])
        ws.add_chart(chart)


def _render_customer_charts(ws, data: _DataSheet, df: pd.DataFrame) -> None:
    from utilities.helpers import bracket_revenue_summary

    _banner(ws, 68, "CUSTOMER ANALYSIS  —  DISTRIBUTION HISTOGRAMS")
    ws.row_dimensions[68].height = 19.5

    if "Customer Count" not in df.columns:
        _merge(ws, 70, FIRST_COL, 70, LAST_COL,
               "Customer histograms unavailable — no Customer Count column in source data.",
               font=_font(10, italic=True, color=GRAY_TEXT), fill=_fill(WHITE),
               align=_align(horizontal="left"))
        return

    histo_specs = [
        ("$ Income", INCOME_BRACKETS, "Customer Count by Income", BLUE),
        ("$ Home Value", HOME_VALUE_BRACKETS, "Customer Count by Home Value", ORANGE),
    ]
    for value_col, brackets, title, color in histo_specs:
        if value_col not in df.columns:
            continue
        summary = bracket_revenue_summary(df, value_col, "Customer Count", brackets)
        labels = [_dollar_bracket(b) for b in summary["Bracket"]]
        cats, vals = data.block(title, labels,
                                list(summary["Total Revenue"]),
                                value_fmt=INT_FMT)
        chart = _bar_chart(title, cats, vals, color, horizontal=False,
                           gap_width=REVENUE_GAP_WIDTH, value_fmt=INT_FMT)
        chart.anchor = _two_cell_anchor(CUSTOMER_CHART_ANCHORS[value_col])
        ws.add_chart(chart)


def _render_profile_summary(ws, base_summary_dict: dict, profile_mode=None,
                            profile_driver=None) -> None:
    prof_c1, prof_c2 = PROFILE_PANEL_COLS
    expl_c1, expl_c2 = EXPLANATION_PANEL_COLS
    lab_c1, lab_c2 = PROFILE_LABEL_COLS
    val_c1, val_c2 = PROFILE_VALUE_COLS

    _banner(ws, SUMMARY_BANNER_ROW, "IDEAL PROFILE SUMMARY", c1=prof_c1, c2=prof_c2)
    _banner(ws, SUMMARY_BANNER_ROW, "EXPLANATION", c1=expl_c1, c2=expl_c2)
    ws.row_dimensions[SUMMARY_BANNER_ROW].height = 19.5

    rows = list(zip(base_summary_dict.get("Note", []),
                    base_summary_dict.get("Value", [])))
    if not rows:
        rows = [("No ideal-profile criteria selected.", "")]
    rows = rows[:SUMMARY_MAX_ROW - SUMMARY_FIRST_ROW + 1]
    last_row = SUMMARY_FIRST_ROW + len(rows) - 1

    _merge(ws, SUMMARY_FIRST_ROW, expl_c1, last_row, expl_c2,
           build_profile_explanation(profile_mode, profile_driver),
           font=_font(10, italic=True, color=GRAY_TEXT), fill=_fill(WHITE),
           align=_align(horizontal="left", vertical="top", wrap=True))
    _outline(ws, SUMMARY_FIRST_ROW, expl_c1, last_row, expl_c2)

    for idx, (note, value) in enumerate(rows):
        row = SUMMARY_FIRST_ROW + idx
        ws.row_dimensions[row].height = 18.0 if idx == 0 else 16.5
        if idx == 0:
            _merge(ws, row, lab_c1, row, lab_c2, note,
                   font=_font(9, bold=True, color=WHITE), fill=_fill(BLUE),
                   align=_align())
            _merge(ws, row, val_c1, row, val_c2, value,
                   font=_font(9, bold=True, color=WHITE), fill=_fill(BLUE),
                   align=_align())
        else:
            band = ALT_ROW if idx % 2 == 1 else WHITE
            _merge(ws, row, lab_c1, row, lab_c2, note,
                   font=_font(10, color=NAVY), fill=_fill(band), align=_align())
            _merge(ws, row, val_c1, row, val_c2, value,
                   font=_font(10, bold=True, color=NAVY), fill=_fill(band),
                   align=_align())
            _outline(ws, row, prof_c1, row, prof_c2, color=BORDER_GRAY)


def _size_columns(ws) -> None:
    ws.column_dimensions["A"].width = GUTTER_COL_WIDTH
    ws.column_dimensions["B"].width = FIRST_CONTENT_WIDTH
    for col in range(3, LAST_COL + 1):  # C–S
        ws.column_dimensions[get_column_letter(col)].width = CONTENT_COL_WIDTH


def _setup_print(ws) -> None:
    ws.print_area = (f"{get_column_letter(FIRST_COL)}1:"
                     f"{get_column_letter(LAST_COL)}{PRINT_LAST_ROW}")
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = PAPER_LETTER
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(**PRINT_MARGINS)
    ws.print_options.gridLines = False
    ws.print_options.horizontalCentered = True
