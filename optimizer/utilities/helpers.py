import math
import re

import numpy as np
import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter

from utilities.config import PRIVILEGED_USERS

# --- User Helpers ---

def normalize_user_role(auth):
    if not auth:
        return ""
    return str(auth).strip().lower()

def is_admin(auth):
    return normalize_user_role(auth) in {role.strip().lower() for role in PRIVILEGED_USERS}

def can_access(auth, feature_name, permissions_map):
    allowed_roles = permissions_map.get(feature_name)

    if allowed_roles is None:
        return True

    user_role = normalize_user_role(auth)
    normalized_allowed_roles = {role.strip().lower() for role in allowed_roles}
    return user_role in normalized_allowed_roles


# --- Data Manipulation Helpers ---

def to_snake_case_list(strings: list[str]) -> list[str]:
    def to_snake(s: str) -> str:
        s = s.lower()
        s = re.sub(r'[^a-z0-9]+', ' ', s)
        return '_'.join(s.strip().split())
    
    return [to_snake(s) for s in strings]

def unsnake(s: str) -> str:
    parts = s.split('_')
    return ' '.join(word.capitalize() for word in parts)

def coerce_numeric_columns(frame: pd.DataFrame, columns: set[str]) -> None:
    for col in columns:
        if col not in frame.columns:
            continue
        series = frame[col]
        if pd.api.types.is_numeric_dtype(series):
            frame[col] = pd.to_numeric(series, errors="coerce")
            continue
        cleaned = series.astype(str).str.replace(r"[,\$%]", "", regex=True)
        frame[col] = pd.to_numeric(cleaned, errors="coerce")

def coerce_year_built(series: pd.Series, min_year: int = 1800, max_year: int | None = None) -> pd.Series:
    if max_year is None:
        max_year = pd.Timestamp.today().year + 1

    num = pd.to_numeric(series, errors="coerce")
    out = pd.Series(np.nan, index=series.index, dtype="float64")

    # Already a year
    direct_year = num.between(min_year, max_year)
    out.loc[direct_year] = num.loc[direct_year]

    # Excel serial date -> year
    serial = out.isna() & num.between(1, 60000)
    out.loc[serial] = pd.to_datetime(
        num.loc[serial], unit="D", origin="1899-12-30", errors="coerce"
    ).dt.year

    # Datetime/string fallback
    rem = out.isna()
    out.loc[rem] = pd.to_datetime(series.loc[rem], errors="coerce").dt.year

    out = out.where(out.between(min_year, max_year))
    return out.round().astype("Int64")

def sanitize_ideal_year(value, min_year: int = 1800, max_year: int | None = None):
    if max_year is None:
        max_year = pd.Timestamp.today().year + 1
    if pd.isna(value):
        return None
    y = int(round(float(value)))
    return y if min_year <= y <= max_year else None

def normalize(df, weight_keys, columns, audit, id_series):
    predictors_to_normalize = set()
    predictors_to_normalize |= set([k for k in weight_keys if k in columns])
    predictors_to_normalize.add("Distance")
    
    for col in predictors_to_normalize:
        try:
            norm, col_audit = adaptive_minmax_iqr(df[col], col_name=col, id_series=id_series)
            df[f"{col}_Norm"] = norm
            audit.append(col_audit)
        except Exception:
            pass  
    return df, audit

def adaptive_minmax_iqr(s: pd.Series, col_name: str, id_series: pd.Series | None = None):
    s = pd.to_numeric(s, errors='coerce')
    s_nonnull = s.dropna()

    audit = {
        "column": col_name,
        "has_outliers": False,
        "lower_fence": np.nan,
        "upper_fence": np.nan,
        "num_clipped": 0,
        "clipped_row_indices": [],
        "clipped_row_ids": [],
        "clipped_original_values": [],
        "clipped_new_values": [],
    }

    if s_nonnull.empty:
        return pd.Series(0.0, index=s.index), audit

    q1, q3 = s_nonnull.quantile([0.25, 0.75])
    iqr = q3 - q1
    if iqr > 0:
        lower, upper = q1 - 1.5*iqr, q3 + 1.5*iqr
        audit["lower_fence"] = float(lower)
        audit["upper_fence"] = float(upper)

        mask_low = s_nonnull < lower
        mask_high = s_nonnull > upper
        has_out = mask_low.any() or mask_high.any()
        audit["has_outliers"] = bool(has_out)

        if has_out:
            clipped_idx = s.index[(s < lower) | (s > upper)]
            audit["clipped_row_indices"] = clipped_idx.tolist()
            audit["num_clipped"] = int(len(clipped_idx))
            audit["clipped_original_values"] = s.loc[clipped_idx].tolist()
            audit["clipped_new_values"] = s.loc[clipped_idx].clip(lower, upper).tolist()
            if id_series is not None:
                audit["clipped_row_ids"] = id_series.loc[clipped_idx].astype(str).tolist()

            s_clip = s.clip(lower, upper)
        else:
            s_clip = s
    else:
        s_clip = s  # constant column (no spread)

    mn, mx = s_clip.min(), s_clip.max()
    if pd.isna(mn) or pd.isna(mx) or mx <= mn:
        return pd.Series(0.0, index=s.index), audit

    norm = ((s_clip - mn) / (mx - mn)).reindex(s.index).fillna(0.0)
    return norm, audit

# --- Customer Profile Helpers ---

def top_quartile_subset(frame: pd.DataFrame, by_col: str) -> pd.DataFrame:
    series = pd.to_numeric(frame[by_col], errors="coerce")
    q = series.quantile(0.75)
    return frame[series >= q]

def build_profiles_df(frame, available_drivers, ideal_map):
    profiles = {}
    for drv in available_drivers:
        top = top_quartile_subset(frame, by_col=drv)
        vals = {}

        for col, ideal_label in ideal_map.items():
            if col not in top.columns:
                vals[ideal_label] = None
                continue

            if col == "Median Year Structure Built":
                y = coerce_year_built(top[col])
                vals[ideal_label] = sanitize_ideal_year(y.median())
            else:
                vals[ideal_label] = top[col].median()
        
        profiles[drv] = vals

    profiles_df = pd.DataFrame(profiles)
    profiles_df.index.name = "Ideal"
    return profiles_df

def extract_driver_ideals(profiles_df, driver, ideal_map):
    return {
        col: (
            profiles_df.loc[ideal_label, driver]
            if ideal_label in profiles_df.index and driver in profiles_df.columns
            else None
        )
        for col, ideal_label in ideal_map.items()
    }

# --- Configuration Helpers ---

def feature(category, name, category_config):
    return category_config[category]["features"].get(name, False)

def get_category_presets(category, category_config):
    return category_config[category]["presets"]

def get_export_layout(category, category_config):
    return category_config[category].get("export_layout", {})

def get_default_weights(category, preset, category_config):
    presets = get_category_presets(category, category_config)
    w = presets.get(preset, {})
    if w:
        return w
    return category_config[category]["fallback_weights"]

def required_columns_for(category, preset, category_config):
    base = category_config[category]["required_columns_base"]
    extra = category_config[category]["required_columns_by_preset"].get(preset, [])
    return list(dict.fromkeys(base + extra))  # preserve order, remove duplicates

# --- Weight Helpers ---

def normalize_weight_defaults(weights, target: float = 1.0, tol: float = 1e-9):
    clean = {k: max(0.0, float(v)) for k, v in weights.items()}
    total = sum(clean.values())
    if total <= tol:
        n = len(clean)
        return {k: (target / n if n else 0.0) for k in clean}
    scale = target / total
    return {k: v * scale for k, v in clean.items()} 

def rebalance_unlocked_weights(
    weights,
    locked_flags,
    target: float = 1.0,
    min_weight: float = 0.0,
    max_weight: float = 1.0,
    tol: float = 1e-9,
):
    if min_weight > max_weight:
        out = {k: float(v) for k, v in weights.items()}
        return out, f"Invalid bounds: min_weight ({min_weight}) > max_weight ({max_weight})."

    def clamp(x):
        return min(max(float(x), min_weight), max_weight)

    out = {k: clamp(v) for k, v in weights.items()}
    locked_keys = {k for k, is_locked in locked_flags.items() if is_locked and k in out}
    unlocked_keys = [k for k in out if k not in locked_keys]

    locked_sum = sum(out[k] for k in locked_keys)
    if locked_sum > target + tol:
        return out, f"Locked weights already sum to {locked_sum:.3f}, above target {target:.3f}."

    if not unlocked_keys:
        total = sum(out.values())
        if abs(total - target) > tol:
            return out, f"All weights are locked and total is {total:.3f}, not {target:.3f}."
        return out, None

    # Feasibility check for unlocked weights under bounds.
    min_possible = locked_sum + len(unlocked_keys) * min_weight
    max_possible = locked_sum + len(unlocked_keys) * max_weight
    if target < min_possible - tol or target > max_possible + tol:
        return out, (
            f"Target {target:.3f} is infeasible with current locks/bounds. "
            f"Feasible range is [{min_possible:.3f}, {max_possible:.3f}]."
        )

    remaining = target - locked_sum

    # Start from proportional scaling of current unlocked weights.
    unlocked_sum = sum(out[k] for k in unlocked_keys)
    if unlocked_sum <= tol:
        even = remaining / len(unlocked_keys)
        for k in unlocked_keys:
            out[k] = clamp(even)
    else:
        scale = remaining / unlocked_sum
        for k in unlocked_keys:
            out[k] = clamp(out[k] * scale)

    # Iteratively redistribute residual across unlocked keys with available room.
    max_iter = len(unlocked_keys) * 8 + 8
    for _ in range(max_iter):
        current_unlocked_sum = sum(out[k] for k in unlocked_keys)
        diff = remaining - current_unlocked_sum
        if abs(diff) <= tol:
            break

        if diff > 0:
            movable = [k for k in unlocked_keys if out[k] < max_weight - tol]
        else:
            movable = [k for k in unlocked_keys if out[k] > min_weight + tol]

        if not movable:
            break

        delta = diff / len(movable)
        for k in movable:
            out[k] = clamp(out[k] + delta)

    final_sum = sum(out.values())
    if abs(final_sum - target) > tol:
        return out, f"After rebalancing, total weights sum to {final_sum:.3f}, not {target:.3f}."

    return out, None

def snap_weights_to_step_preserve_sum(
    weights,
    locked_flags,
    target: float = 1.0,
    step: float = 0.01,
    min_weight: float = 0.0,
    max_weight: float = 1.0,
    tol: float = 1e-9,
):
    if step <= 0:
        return dict(weights), f"Invalid step: {step}. Must be > 0."
    if min_weight > max_weight:
        return dict(weights), f"Invalid bounds: min_weight ({min_weight}) > max_weight ({max_weight})."

    def clamp(x):
        return min(max(float(x), min_weight), max_weight)

    out = {k: clamp(v) for k, v in weights.items()}
    locked_keys = {k for k, is_locked in locked_flags.items() if is_locked and k in out}
    unlocked_keys = [k for k in out if k not in locked_keys]

    locked_sum = sum(out[k] for k in locked_keys)
    remaining = target - locked_sum

    if not unlocked_keys:
        total = sum(out.values())
        if abs(total - target) > tol:
            return out, f"All weights are locked and total is {total:.3f}, not {target:.3f}."
        return out, None

    n = len(unlocked_keys)

    min_possible = n * min_weight
    max_possible = n * max_weight
    if remaining < min_possible - tol or remaining > max_possible + tol:
        return out, (
            f"Target {target:.3f} is infeasible with locks/bounds. "
            f"Unlocked contribution must be in [{min_possible:.3f}, {max_possible:.3f}], got {remaining:.3f}."
        )
    
    span = max_weight - min_weight
    max_units = int(math.floor((span / step) + tol))
    if max_units < 0:
        return out, "Invalid max units computed from bounds and step."

    target_units_float = (remaining - n * min_weight) / step
    target_units = int(round(target_units_float))

    if abs(target_units_float - target_units) > 1e-6:
        return out, (
            f"Cannot exactly hit target with step {step}. "
            f"Remaining unlocked total {remaining:.6f} is not step-aligned."
        )

    units = {}
    frac = {}

    for k in unlocked_keys:
        u_float = (out[k] - min_weight) / step
        u_float = min(max(u_float, 0.0), float(max_units))
        u_floor = int(math.floor(u_float + tol))
        units[k] = u_floor
        frac[k] = u_float - u_floor

    current_units = sum(units.values())
    diff_units = target_units - current_units

    while diff_units > 0:
        candidates = [k for k in unlocked_keys if units[k] < max_units]
        if not candidates:
            return out, "Unable to distribute positive residual within bounds."
        k = max(candidates, key=lambda kk: (frac[kk], max_units - units[kk]))
        units[k] += 1
        diff_units -= 1

    
    while diff_units < 0:
        candidates = [k for k in unlocked_keys if units[k] > 0]
        if not candidates:
            return out, "Unable to remove residual within bounds."
        k = min(candidates, key=lambda kk: (frac[kk], units[kk]))
        units[k] -= 1
        diff_units += 1

    for k in unlocked_keys:
        out[k] = clamp(min_weight + units[k] * step)

    final_sum = sum(out.values())
    if abs(final_sum - target) > 1e-6:
        return out, f"After step-snap, total is {final_sum:.6f}, not {target:.6f}."

    return out, None


# --- Output Helpers ---

def bracket_revenue_summary(
    frame: pd.DataFrame,
    value_col: str,
    revenue_col: str,
    brackets: list[tuple[int, int | None, str]],
) -> pd.DataFrame:
    values = pd.to_numeric(frame[value_col], errors="coerce")
    revenue = pd.to_numeric(frame[revenue_col], errors="coerce").fillna(0)

    rows = []
    for lower, upper, label in brackets:
        if upper is None:
            mask = values >= lower
        else:
            mask = (values >= lower) & (values <= upper)
        rows.append(
            {
                "Bracket": label,
                "Total Revenue": float(revenue[mask].sum()),
                "Count": int(mask.sum()),
            }
        )

    return pd.DataFrame(rows)

def add_revenue_histograms_sheet(
    writer: pd.ExcelWriter,
    working_df: pd.DataFrame,
    revenue_col: str,
):
    income_brackets = [
        (0, 39_999, "0-39,999"),
        (40_000, 74_999, "40,000-74,999"),
        (75_000, 149_999, "75,000-149,999"),
        (150_000, 174_999, "150,000-174,999"),
        (175_000, 249_999, "175,000-249,999"),
        (250_000, None, "250,000+"),
    ]
    home_value_brackets = [
        (0, 149_999, "0-149,999"),
        (150_000, 249_999, "150,000-249,999"),
        (250_000, 399_999, "250,000-399,999"),
        (400_000, 749_999, "400,000-749,999"),
        (750_000, 999_999, "750,000-999,999"),
        (1_000_000, None, "1,000,000+"),
    ]

    sheet_name = "Revenue Histograms"
    income_df = bracket_revenue_summary(working_df, "$ Income", revenue_col, income_brackets)
    home_value_df = bracket_revenue_summary(working_df, "$ Home Value", revenue_col, home_value_brackets)

    income_startrow = 0
    income_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=income_startrow)

    home_value_startrow = income_startrow + len(income_df) + 4
    home_value_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=home_value_startrow)

    ws = writer.sheets[sheet_name]

    income_chart = BarChart()
    income_chart.type = "col"
    income_chart.title = "Revenue by Income Bracket"
    income_chart.y_axis.title = "Total Revenue"
    income_chart.x_axis.title = "Income Bracket"
    income_data = Reference(
        ws,
        min_col=2,
        min_row=income_startrow + 1,
        max_row=income_startrow + len(income_df) + 1,
    )
    income_cats = Reference(
        ws,
        min_col=1,
        min_row=income_startrow + 2,
        max_row=income_startrow + len(income_df) + 1,
    )
    income_chart.add_data(income_data, titles_from_data=True)
    income_chart.set_categories(income_cats)
    income_chart.height = 12
    income_chart.width = 20
    ws.add_chart(income_chart, "D2")

    home_chart = BarChart()
    home_chart.type = "col"
    home_chart.title = "Revenue by Home Value Bracket"
    home_chart.y_axis.title = "Total Revenue"
    home_chart.x_axis.title = "Home Value Bracket"
    home_data = Reference(
        ws,
        min_col=2,
        min_row=home_value_startrow + 1,
        max_row=home_value_startrow + len(home_value_df) + 1,
    )
    home_cats = Reference(
        ws,
        min_col=1,
        min_row=home_value_startrow + 2,
        max_row=home_value_startrow + len(home_value_df) + 1,
    )
    home_chart.add_data(home_data, titles_from_data=True)
    home_chart.set_categories(home_cats)
    home_chart.height = 12
    home_chart.width = 20
    ws.add_chart(home_chart, f"D{home_value_startrow + 2}")

def flag_and_filter(df, max_penetration, max_income, min_income, max_distance, min_owner):
    if max_penetration is not None and 'House Penetration%' in df.columns:
        df['Penetration Flag'] = np.where(
            df['House Penetration%'] > max_penetration,
            "⚠️ Above Max Penetration", ""
        )
    df['Status'] = np.where(
        ((max_income is None) | (df['$ Income'] <= max_income)) &
        ((min_income is None) | (df['$ Income'] >= min_income)) &
        ((max_distance is None) | (df['Distance'] <= max_distance)) &
        ((min_owner is None) | (df['Owner Occupied'] >= min_owner)),
        "Included", "Excluded"
    )
    return df

def create_summary(audit, profile, ideal_values, ideal_map):
        # Summary per column
    audit_summary_df = pd.DataFrame([{
        "Column": a["column"],
        "Outliers Detected": a["has_outliers"],
        "Lower Fence": a["lower_fence"],
        "Upper Fence": a["upper_fence"],
        "Rows Winsorized (count)": a["num_clipped"],
        "Winsorized Geocodes (preview)": ", ".join(a["clipped_row_ids"][:10]) if a["clipped_row_ids"] else ""
    } for a in audit])

    detail_rows = []
    for a in audit:
        if a.get("has_outliers") and a.get("clipped_row_indices"):
            for rid, orig, new in zip(
                a.get("clipped_row_ids", []),
                a.get("clipped_original_values", []),
                a.get("clipped_new_values", [])
            ):
                detail_rows.append({
                    "Column": a.get("column"),
                    "Geocode": rid,
                    "Row Index": a.get("clipped_row_indices", [])[a.get("clipped_row_ids", []).index(rid)] if a.get("clipped_row_ids", []) else None,
                    "Original Value": orig,
                    "Clipped Value": new
                })
    winsor_details_df = pd.DataFrame(detail_rows)

    # Base summary 
    base_summary = {
        "Note": [],
        "Value": [],
    }
    if profile:
        base_summary["Note"].append("Profile Mode Used:")
        base_summary["Value"].append(profile)
    
    for col, ideal in ideal_values.items():
        if ideal is None:
            continue

        if col == "Median Year Structure Built":
            ideal = sanitize_ideal_year(ideal)
            if ideal is None:
                continue

        base_summary["Note"].append(f"{ideal_map[col]}:")
        base_summary["Value"].append(ideal)

    return audit_summary_df, winsor_details_df, base_summary

def prepare_ranked_columns(df, drop, order, keep_only_ordered):
    out = df.copy()
    if drop:
        out = out.drop(columns=drop, errors="ignore")
    if not order:
        return out
    
    ordered = [col for col in order if col in out.columns]

    if keep_only_ordered:
        return out.reindex(columns=ordered)

    remaining = [col for col in out.columns if col not in ordered]
    return out.reindex(columns=ordered + remaining)

def style_header_row(ws, header_colors, font_color="FF000000"):
    if not header_colors:
        return
    
    for cell in ws[1]:
        color = header_colors.get(cell.value)
        if not color:
            continue

        hex_color = color.lstrip('#').upper()
        if len(hex_color) == 6:
            hex_color = 'FF' + hex_color  # add full opacity if not provided

        cell.fill = PatternFill(fill_type="solid", fgColor=hex_color)
        cell.font = Font(bold=True, color=font_color)

def highlight_excluded_rows(ws, status_header="Status"):
    # Find the Status column in the header row
    status_cell = next((cell for cell in ws[1] if cell.value == status_header), None)
    if status_cell is None:
        return

    if ws.max_row < 2 or ws.max_column < 1:
        return

    status_col_letter = get_column_letter(status_cell.column)
    first_data_row = 2
    last_row = ws.max_row
    last_col_letter = get_column_letter(ws.max_column)

    red_fill = PatternFill(fill_type="solid", fgColor="FFFFC7CE")
    red_font = Font(color="FF9C0006")
    dxf = DifferentialStyle(fill=red_fill, font=red_font)

    formula = f'${status_col_letter}{first_data_row}="Excluded"'
    data_range = f"A{first_data_row}:{last_col_letter}{last_row}"

    rule = Rule(type="expression", dxf=dxf, formula=[formula])
    ws.conditional_formatting.add(data_range, rule)


def write_ranked_workbook(
    writer: pd.ExcelWriter,
    working_df: pd.DataFrame,
    base_summary_dict: dict,
    audit_summary_df: pd.DataFrame,
    profiles_df: pd.DataFrame,
    winsor_details_df: pd.DataFrame,
    profile_mode: str | None,
    ranked_columns_order: list[str] | None,
    ranked_columns_drop: list[str] | None,
    ranked_header_colors: dict[str, str] | None,
    include_winsor_sheet: bool,
    ranked_sheet_name: str = "Ranked Geocodes",
    ranked_df_override: pd.DataFrame | None = None,
    report_label: str | None = None,
    profile_driver: str | None = None,
):
    ranked_source_df = ranked_df_override if ranked_df_override is not None else working_df


    # Ranked data
    ranked_df = prepare_ranked_columns(
        df=ranked_source_df,
        drop=ranked_columns_drop,
        order=ranked_columns_order,
        keep_only_ordered=True
    )
    ranked_df.to_excel(writer, sheet_name=ranked_sheet_name, index=False)

    ranked_ws = writer.sheets[ranked_sheet_name]
    style_header_row(ranked_ws, ranked_header_colors)
    highlight_excluded_rows(ranked_ws, status_header="Status")

    # Base summary at the top of 'Summary'
    pd.DataFrame(base_summary_dict).to_excel(writer, sheet_name='Summary', index=False, startrow=0)

    # Winsor summary below it with a blank row
    startrow = (len(base_summary_dict.get("Note", [])) or 1) + 2
    if not audit_summary_df.empty:
        audit_summary_df.to_excel(writer, sheet_name='Summary', index=False, startrow=startrow)

    # Profiles (optional)
    if profile_mode == "Dynamic (from file)" and not profiles_df.empty:
        profiles_df.to_excel(writer, sheet_name='Summary', index=True, startrow=startrow + len(audit_summary_df) + 2)

    # Detailed sheet (optional)
    if include_winsor_sheet and not winsor_details_df.empty:
        winsor_details_df.to_excel(writer, sheet_name='Winsorized Rows', index=False)

    revenue_col = None
    if "$ Total Spend" in working_df.columns:
        revenue_col = "$ Total Spend"
    elif "Overall Revenue" in working_df.columns:
        revenue_col = "Overall Revenue"
    elif "Revenue" in working_df.columns:
        revenue_col = "Revenue"

    if revenue_col and "$ Income" in working_df.columns and "$ Home Value" in working_df.columns:
        add_revenue_histograms_sheet(writer, working_df, revenue_col)
    else:
        sheet_name = "Revenue Histograms"
        pd.DataFrame(
            {
                "Note": [
                    "Revenue histograms not generated. Missing $ Income, $ Home Value, or revenue column."
                ]
            }
        ).to_excel(writer, sheet_name=sheet_name, index=False)

    # Imported lazily to avoid an import cycle (dashboard imports from this file).
    from utilities.dashboard import add_dashboard_sheet

    add_dashboard_sheet(
        writer,
        working_df=ranked_source_df,
        base_summary_dict=base_summary_dict,
        report_label=report_label,
        profile_mode=profile_mode,
        profile_driver=profile_driver,
    )


def restrict_visible_sheets(workbook, visible_sheets):
    keep = set(visible_sheets)
    first_visible = None
    for ws in workbook.worksheets:
        if ws.title in keep:
            ws.sheet_state = "visible"
            first_visible = first_visible or ws
        else:
            ws.sheet_state = "hidden"
    if first_visible is not None:
        workbook.active = workbook.index(first_visible)
